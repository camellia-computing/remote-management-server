import csv
from contextlib import suppress
from datetime import UTC, datetime
from io import BytesIO
from tempfile import SpooledTemporaryFile
from time import monotonic
from zipfile import ZIP_DEFLATED, ZipFile

from django.http import FileResponse, HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class SpreadsheetBudgetExceeded(ValueError):
    pass


class _BudgetedWorkbookTarget:
    def __init__(self, target, *, max_bytes, deadline):
        self.target = target
        self.max_bytes = max_bytes
        self.deadline = deadline
        self.bytes_written = 0
        self.enforce_budget = True

    def write(self, data):
        if self.enforce_budget:
            if monotonic() > self.deadline:
                raise SpreadsheetBudgetExceeded("spreadsheet deadline exceeded")
            next_size = self.bytes_written + len(data)
            if next_size > self.max_bytes:
                raise SpreadsheetBudgetExceeded("spreadsheet byte budget exceeded")
        written = self.target.write(data)
        self.bytes_written += written
        return written

    def __getattr__(self, name):
        return getattr(self.target, name)


def _save_budgeted_workbook(workbook, target):
    archive = ZipFile(target, "w", ZIP_DEFLATED, allowZip64=True)
    workbook.properties.modified = datetime.now(tz=UTC).replace(tzinfo=None)
    try:
        ExcelWriter(workbook, archive).save()
    except Exception:
        # openpyxl closes its ZipFile only after a successful write. Disable the
        # response budget solely while abort-closing the doomed temporary file,
        # then let the caller delete it instead of leaving a live ZipFile finalizer.
        target.enforce_budget = False
        with suppress(Exception):
            archive.close()
        raise


def spreadsheet_safe_value(value):
    if value is None:
        return ""
    if not isinstance(value, str) or not value:
        return value
    trimmed = value.lstrip(" \t\r\n")
    if value[0] in "\t\r\n" or (trimmed and trimmed[0] in "=+-@"):
        return "'" + value
    return value


class _SafeCsvWriter:
    def __init__(self, output, **kwargs):
        self._writer = csv.writer(output, **kwargs)

    def writerow(self, row):
        return self._writer.writerow([spreadsheet_safe_value(value) for value in row])

    def writerows(self, rows):
        return self._writer.writerows([spreadsheet_safe_value(value) for value in row] for row in rows)


def safe_csv_writer(output, **kwargs):
    return _SafeCsvWriter(output, **kwargs)


def xlsx_response(filename, sheet_name, headers, rows):
    workbook = Workbook()
    sheet = workbook.active
    safe_sheet_name = "".join("_" if character in r"[]:*?/\\" else character for character in str(sheet_name))
    sheet.title = safe_sheet_name[:31] or "Sheet1"
    sheet.append([spreadsheet_safe_value(value) for value in headers])
    for row in rows:
        sheet.append([spreadsheet_safe_value(value) for value in row])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type=XLSX_CONTENT_TYPE)
    response["Content-Disposition"] = f"attachment; filename={filename}"
    return response


def bounded_xlsx_file_response(
    filename,
    sheet_name,
    headers,
    rows,
    *,
    max_rows,
    max_bytes,
    deadline_seconds,
    spool_memory_bytes,
):
    deadline = monotonic() + deadline_seconds
    target = SpooledTemporaryFile(max_size=spool_memory_bytes, mode="w+b")
    workbook = Workbook(write_only=True)
    sheet = None
    try:
        sheet = workbook.create_sheet()
        safe_sheet_name = "".join("_" if character in r"[]:*?/\\" else character for character in str(sheet_name))
        sheet.title = safe_sheet_name[:31] or "Sheet1"
        sheet.append([spreadsheet_safe_value(value) for value in headers])
        row_count = 0
        for row in rows:
            if row_count >= max_rows:
                raise SpreadsheetBudgetExceeded("spreadsheet row budget exceeded")
            if monotonic() > deadline:
                raise SpreadsheetBudgetExceeded("spreadsheet deadline exceeded")
            sheet.append([spreadsheet_safe_value(value) for value in row])
            row_count += 1

        budgeted_target = _BudgetedWorkbookTarget(
            target,
            max_bytes=max_bytes,
            deadline=deadline,
        )
        _save_budgeted_workbook(workbook, budgeted_target)
        if monotonic() > deadline:
            raise SpreadsheetBudgetExceeded("spreadsheet deadline exceeded")
        response_bytes = target.tell()
        target.seek(0)
        response = FileResponse(
            target,
            as_attachment=True,
            filename=filename,
            content_type=XLSX_CONTENT_TYPE,
        )
        response["Content-Length"] = str(response_bytes)
        return response, row_count, response_bytes
    except Exception:
        if sheet is not None and not sheet.closed:
            with suppress(Exception):
                sheet.close()
        target.close()
        raise
    finally:
        workbook.close()
