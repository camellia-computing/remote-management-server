import base64
import datetime
import hashlib
import json
import os
import tempfile
import uuid
from io import BytesIO, StringIO
from pathlib import Path
from unittest.mock import patch

from django.contrib.auth.models import Group
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import connection, transaction
from django.test import SimpleTestCase, TestCase, override_settings
from django.test.utils import CaptureQueriesContext
from django.utils import timezone
from nacl.secret import SecretBox
from nacl.signing import SigningKey
from openpyxl import load_workbook

from api import ingestion_governance, recording_crypto
from api.encrypted_fields import FIELD_PREFIX, encrypt_text, key_canary, key_fingerprint
from api.formatting import format_bytes
from api.models import (
    AddressBookProfile,
    AddressBookRule,
    AlarmLog,
    ConnLog,
    DataEncryptionKeyState,
    DeviceGroup,
    FileLog,
    LoginAdmissionLock,
    LoginAttempt,
    OidcPendingAuth,
    RecordingUpload,
    RemoteDevice,
    RemotePeer,
    RemoteTag,
    RemoteToken,
    RequestRateBucket,
    RequestRateLease,
    ShareLink,
    StrategyProfile,
    UserProfile,
)
from api.tag_colors import normalize_tag_color, tag_color_css
from api.xlsx import spreadsheet_safe_value
from camellia_remote_management import settings as project_settings
from camellia_remote_management.settings import (
    _database_host,
    _database_text,
    _database_tls_options,
    data_encryption_key_id,
    data_encryption_legacy_keys,
    env_bool,
    env_choice,
    env_int,
)


def device_uuid(label):
    return base64.b64encode(label.encode()).decode()


DEFAULT_DEVICE_UUID = device_uuid("device-uuid")


class UtilityContractTests(SimpleTestCase):
    def test_byte_formatting_uses_binary_units_and_rejects_invalid_values(self):
        self.assertEqual(format_bytes(0), "0 B")
        self.assertEqual(format_bytes(1024), "1 KiB")
        self.assertEqual(format_bytes(1536), "1.5 KiB")
        with self.assertRaises(TypeError):
            format_bytes(True)
        with self.assertRaises(ValueError):
            format_bytes(-1)

    def test_tag_colors_have_one_bounded_argb_representation(self):
        self.assertEqual(normalize_tag_color("#336699"), str(0xFF336699))
        self.assertEqual(normalize_tag_color("#80336699"), str(0x80336699))
        self.assertEqual(tag_color_css(str(0xFF336699)), "#336699")
        for value in (True, -1, 0x1_0000_0000, "rgb(1,2,3)", "#12345"):
            with self.subTest(value=value):
                self.assertIsNone(normalize_tag_color(value))

    def test_spreadsheet_exports_neutralize_formula_prefixes(self):
        for value in ("=1+1", " +SUM(A1:A2)", "\t@command", "-2+3"):
            with self.subTest(value=value):
                self.assertTrue(spreadsheet_safe_value(value).startswith("'"))
        self.assertEqual(spreadsheet_safe_value("ordinary text"), "ordinary text")
        self.assertEqual(spreadsheet_safe_value(42), 42)


class SettingsParserTests(SimpleTestCase):
    def test_data_encryption_key_ids_and_legacy_keyring_are_strict(self):
        first_key = base64.b64encode(b"a" * 32).decode()
        second_key = base64.b64encode(b"b" * 32).decode()
        self.assertEqual(data_encryption_key_id("key-2026.08", "TEST_KEY_ID"), "key-2026.08")
        self.assertEqual(
            data_encryption_legacy_keys(f"old-a:{first_key},old-b:{second_key}"),
            {"old-a": b"a" * 32, "old-b": b"b" * 32},
        )
        for value in ("", "UPPERCASE", "-leading", "x" * 33, "contains space"):
            with self.subTest(value=value), self.assertRaises(ImproperlyConfigured):
                data_encryption_key_id(value, "TEST_KEY_ID")
        for value in (
            "missing-separator",
            "duplicate:" + first_key + ",duplicate:" + second_key,
            "first:" + first_key + ",alias:" + first_key,
            "invalid:not-base64",
        ):
            with self.subTest(value=value), self.assertRaises(ImproperlyConfigured):
                data_encryption_legacy_keys(value)

    def test_boolean_values_are_explicit_and_case_insensitive(self):
        with patch.dict(os.environ, {"TEST_BOOLEAN": "YeS"}):
            self.assertTrue(env_bool("TEST_BOOLEAN"))
        with patch.dict(os.environ, {"TEST_BOOLEAN": "OFF"}):
            self.assertFalse(env_bool("TEST_BOOLEAN", True))

    def test_invalid_boolean_does_not_silently_disable_security(self):
        with (
            patch.dict(os.environ, {"TEST_BOOLEAN": "treu"}),
            self.assertRaises(ImproperlyConfigured),
        ):
            env_bool("TEST_BOOLEAN")

    def test_invalid_or_out_of_range_integer_does_not_use_a_default(self):
        for value in ("many", "0", "101"):
            with (
                self.subTest(value=value),
                patch.dict(os.environ, {"TEST_INTEGER": value}),
                self.assertRaises(ImproperlyConfigured),
            ):
                env_int("TEST_INTEGER", 10, 1, 100)

    def test_choice_normalization_is_strict(self):
        with patch.dict(os.environ, {"TEST_CHOICE": "debug"}):
            self.assertEqual(env_choice("TEST_CHOICE", "INFO", {"DEBUG", "INFO"}), "DEBUG")
        with (
            patch.dict(os.environ, {"TEST_CHOICE": "verbose"}),
            self.assertRaises(ImproperlyConfigured),
        ):
            env_choice("TEST_CHOICE", "INFO", {"DEBUG", "INFO"})

    def test_database_parameters_preserve_strong_password_characters(self):
        password = "correct horse:@/%# battery staple"
        self.assertEqual(
            _database_text(password, "TEST_DATABASE_PASSWORD", 1024),
            password,
        )
        for host in ("postgres", "db.internal.example", "127.0.0.1", "2001:db8::1"):
            with self.subTest(host=host):
                self.assertEqual(_database_host(host), host)
        for host in ("https://db.example", "db host", "-db.example"):
            with self.subTest(host=host), self.assertRaises(ImproperlyConfigured):
                _database_host(host)

    def test_external_database_tls_fails_closed(self):
        with (
            patch.object(project_settings, "DEBUG", False),
            patch.dict(
                os.environ,
                {
                    "CAMELLIA_REMOTE_DATABASE_SSLMODE": "require",
                    "CAMELLIA_REMOTE_DATABASE_SSLROOTCERT": "",
                    "CAMELLIA_REMOTE_DATABASE_SSLCERT": "",
                    "CAMELLIA_REMOTE_DATABASE_SSLKEY": "",
                },
                clear=False,
            ),
            self.assertRaises(ImproperlyConfigured),
        ):
            _database_tls_options("db.example.test")

        with (
            patch.object(project_settings, "DEBUG", False),
            patch.dict(
                os.environ,
                {
                    "CAMELLIA_REMOTE_DATABASE_SSLMODE": "verify-full",
                    "CAMELLIA_REMOTE_DATABASE_SSLROOTCERT": "/run/secrets/database-ca.pem",
                    "CAMELLIA_REMOTE_DATABASE_SSLCERT": "",
                    "CAMELLIA_REMOTE_DATABASE_SSLKEY": "",
                },
                clear=False,
            ),
        ):
            options = _database_tls_options("db.example.test")
        self.assertEqual(options["sslmode"], "verify-full")
        self.assertEqual(options["sslrootcert"], "/run/secrets/database-ca.pem")

    def test_database_client_certificate_and_key_are_atomic_configuration(self):
        with (
            patch.dict(
                os.environ,
                {
                    "CAMELLIA_REMOTE_DATABASE_SSLMODE": "disable",
                    "CAMELLIA_REMOTE_DATABASE_SSLROOTCERT": "",
                    "CAMELLIA_REMOTE_DATABASE_SSLCERT": "/run/secrets/database-client.pem",
                    "CAMELLIA_REMOTE_DATABASE_SSLKEY": "",
                },
                clear=False,
            ),
            self.assertRaises(ImproperlyConfigured),
        ):
            _database_tls_options("127.0.0.1")


class ApiTestMixin:
    def setUp(self):
        self.admin = UserProfile.objects.create_user(
            username="admin",
            password="admin-pass",
            is_admin=True,
            is_superuser=True,
        )
        self.user = UserProfile.objects.create_user(
            username="alice",
            password="alice-pass",
            email="alice@example.test",
        )
        self.device_signing_keys = {}

    def _post_json(self, path, payload, token=None):
        headers = self._auth_headers(token)
        headers["HTTP_IDEMPOTENCY_KEY"] = str(uuid.uuid4())
        return self.client.post(
            path,
            data=json.dumps(payload),
            content_type="application/json",
            **headers,
        )

    def _put_json(self, path, payload, token=None):
        headers = self._auth_headers(token)
        return self.client.put(
            path,
            data=json.dumps(payload),
            content_type="application/json",
            **headers,
        )

    def _delete_json(self, path, payload=None, token=None):
        headers = self._auth_headers(token)
        return self.client.delete(
            path,
            data=json.dumps(payload if payload is not None else {}),
            content_type="application/json",
            **headers,
        )

    @staticmethod
    def _auth_headers(token):
        return {"HTTP_AUTHORIZATION": f"Bearer {token}"} if token else {}

    def _login(self, username, password, rid="123456789", uuid=DEFAULT_DEVICE_UUID):
        payload = {
            "username": username,
            "password": password,
            "id": rid,
            "uuid": uuid,
            "type": "client",
            "deviceInfo": {
                "os": "linux",
                "type": "client",
                "name": "desktop",
            },
        }
        signing_key = self.device_signing_keys.get((rid, uuid))
        if signing_key is not None:
            payload["device_proof"] = self._device_proof("login", signing_key, rid, uuid)
        response = self._post_json(
            "/api/login",
            payload,
        )
        self.assertEqual(response.status_code, 200, response.content)
        body = response.json()
        self.assertIn("access_token", body)
        return body["access_token"]

    def _device_proof(self, purpose, signing_key, rid, uuid, token=None):
        public_key = base64.b64encode(bytes(signing_key.verify_key)).decode("ascii")
        challenge = self._post_json(
            "/api/devices/proof-challenge",
            {"purpose": purpose, "id": rid, "uuid": uuid, "pk": public_key},
            token=token,
        )
        self.assertEqual(challenge.status_code, 200, challenge.content)
        challenge_body = challenge.json()
        signature = signing_key.sign(challenge_body["message"].encode("utf-8")).signature
        return {
            "challenge": challenge_body["challenge"],
            "public_key": public_key,
            "signature": base64.b64encode(signature).decode("ascii"),
        }

    def _device(self, owner=None, rid="123456789", uuid=DEFAULT_DEVICE_UUID, **overrides):
        signing_key = SigningKey.generate()
        data = {
            "rid": rid,
            "cpu": "-",
            "hostname": "desktop",
            "memory": "-",
            "os": "linux",
            "uuid": uuid,
            "username": "desktop-user",
            "version": "2.0.0",
            "public_key_hash": hashlib.sha256(bytes(signing_key.verify_key)).hexdigest(),
            "owner": owner,
        }
        data.update(overrides)
        device = RemoteDevice.objects.create(**data)
        if "public_key_hash" not in overrides:
            self.device_signing_keys[(rid, uuid)] = signing_key
        return device

    @staticmethod
    def _personal_profile(user):
        profile, _created = AddressBookProfile.objects.get_or_create(
            guid=f"personal-{user.id}",
            defaults={
                "name": "My address book",
                "owner": user,
                "rule": 3,
            },
        )
        return profile


class ApiContractTests(ApiTestMixin, TestCase):
    def test_malformed_json_is_a_client_error(self):
        for payload in (b'{"username":', b"[]", b"null", b'"text"'):
            with self.subTest(payload=payload):
                response = self.client.post(
                    "/api/login",
                    data=payload,
                    content_type="application/json",
                )

                self.assertEqual(response.status_code, 400, response.content)
                self.assertEqual(response.json(), {"error": "Invalid JSON payload"})

    def test_routes_reject_unsupported_methods_with_allow_contracts(self):
        for path, method, allowed in (
            ("/api/users", self.client.put, "GET, POST"),
            ("/api/peers", self.client.post, "GET"),
            ("/api/device-group/accessible", self.client.post, "GET"),
            ("/api/home", self.client.post, "GET"),
            ("/webui2/status", self.client.post, "GET"),
        ):
            with self.subTest(path=path):
                response = method(path, data="{}", content_type="application/json")
                self.assertEqual(response.status_code, 405, response.content)
                self.assertEqual(response.headers["Allow"], allowed)

    def test_login_requires_bearer_token_for_current_user(self):
        token = self._login("alice", "alice-pass")
        device = RemoteDevice.objects.get(rid="123456789")
        self.assertEqual(device.hostname, "desktop")
        self.assertEqual(device.os, "linux")
        stored_token = RemoteToken.objects.get(device=device)
        self.assertEqual(len(stored_token.access_token), 64)
        self.assertIsNotNone(stored_token.expires_at)

        query_token_response = self._post_json("/api/currentUser", {"access_token": token})
        self.assertEqual(query_token_response.status_code, 401)

        response = self._post_json("/api/currentUser", {}, token=token)
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["name"], "alice")

    def test_failed_login_returns_401_with_error_body(self):
        response = self._post_json(
            "/api/login",
            {"username": "alice", "password": "wrong", "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
        )
        self.assertEqual(response.status_code, 401)
        self.assertIn("error", response.json())
        self.assertEqual(LoginAttempt.objects.filter(ip="127.0.0.1").count(), 1)

    def test_login_rejects_noncanonical_device_identity(self):
        malformed_values = (
            "not-base64",
            base64.b64encode(b"devices").decode().rstrip("="),
            "",
        )
        for malformed in malformed_values:
            with self.subTest(uuid=malformed):
                response = self._post_json(
                    "/api/login",
                    {
                        "username": "alice",
                        "password": "alice-pass",
                        "id": "123456789",
                        "uuid": malformed,
                    },
                )
                self.assertEqual(response.status_code, 400, response.content)
        self.assertFalse(RemoteToken.objects.exists())

    def test_repeated_login_failures_lock_out_ip(self):
        for _ in range(10):
            self._post_json(
                "/api/login",
                {"username": "alice", "password": "wrong", "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
            )
        locked = self._post_json(
            "/api/login",
            {"username": "alice", "password": "alice-pass", "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
        )
        self.assertEqual(locked.status_code, 429)
        self.assertIn("error", locked.json())

    def test_successful_login_clears_failures_and_stores_hashed_token(self):
        self._post_json(
            "/api/login",
            {"username": "alice", "password": "wrong", "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
        )
        token = self._login("alice", "alice-pass")
        self.assertEqual(LoginAttempt.objects.count(), 0)

        stored = RemoteToken.objects.get(device__owner=self.user)
        self.assertNotEqual(stored.access_token, token)
        self.assertEqual(stored.access_token, hashlib.sha256(token.encode()).hexdigest())

        # currentUser validates the bearer without making another response-body copy.
        response = self._post_json("/api/currentUser", {}, token=token)
        self.assertEqual(response.status_code, 200, response.content)
        self.assertNotIn("access_token", response.json())

    def test_relogin_rotates_the_stored_token(self):
        first = self._login("alice", "alice-pass")
        second = self._login("alice", "alice-pass")
        self.assertNotEqual(first, second)

        stale = self._post_json("/api/currentUser", {}, token=first)
        self.assertEqual(stale.status_code, 401)
        fresh = self._post_json("/api/currentUser", {}, token=second)
        self.assertEqual(fresh.status_code, 200, fresh.content)

    def test_oidc_auth_query_reads_state_from_database(self):
        # Simulates the callback having completed on a different worker: the
        # pending state lives in the DB, not in process memory.
        poll_code = "test-poll-code"
        OidcPendingAuth.objects.create(
            state="test-state",
            poll_code_hash=hashlib.sha256(poll_code.encode()).hexdigest(),
            provider="example",
            rid="123456789",
            device_uuid=DEFAULT_DEVICE_UUID,
            nonce="test-nonce",
            code_verifier="test-code-verifier",
            status=OidcPendingAuth.STATUS_DONE,
            authenticated_user=self.user,
        )
        self.assertNotEqual(poll_code, "test-state")
        self.assertNotIn(poll_code, OidcPendingAuth.objects.get().poll_code_hash)
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT nonce, code_verifier FROM api_oidcpendingauth WHERE state = %s",
                ["test-state"],
            )
            stored_nonce, stored_verifier = cursor.fetchone()
        envelope_prefix = f"{FIELD_PREFIX}{project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID}:"
        self.assertTrue(stored_nonce.startswith(envelope_prefix))
        self.assertTrue(stored_verifier.startswith(envelope_prefix))
        self.assertNotIn("test-nonce", stored_nonce)
        self.assertNotIn("test-code-verifier", stored_verifier)

        state_is_not_a_poll_secret = self._post_json(
            "/api/oidc/auth-query",
            {"code": "test-state", "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
        )
        self.assertNotIn("access_token", state_is_not_a_poll_secret.json())
        response = self._post_json(
            "/api/oidc/auth-query",
            {"code": poll_code, "id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
        )
        self.assertEqual(response.status_code, 200, response.content)
        self.assertIn("access_token", response.json())
        self.assertNotEqual(response.json()["access_token"], poll_code)
        self.assertFalse(OidcPendingAuth.objects.filter(state="test-state").exists())

    def test_telemetry_requires_the_token_bound_to_that_device(self):
        response = self._post_json(
            "/api/heartbeat",
            {"id": "x" * 200, "uuid": DEFAULT_DEVICE_UUID},
        )
        self.assertEqual(response.status_code, 400)
        self.assertFalse(RemoteDevice.objects.exists())

        self._device(owner=self.user)
        token = self._login("alice", "alice-pass")
        unauthenticated = self._post_json(
            "/api/sysinfo",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID, "hostname": "desktop"},
        )
        self.assertEqual(unauthenticated.status_code, 401)
        wrong_device = self._post_json(
            "/api/sysinfo",
            {"id": "987654321", "uuid": device_uuid("other-device"), "hostname": "desktop"},
            token=token,
        )
        self.assertEqual(wrong_device.status_code, 401)

        sysinfo = self._post_json(
            "/api/sysinfo",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID, "hostname": "desktop"},
            token=token,
        )
        self.assertEqual(sysinfo.status_code, 200, sysinfo.content)
        self.assertTrue(RemoteDevice.objects.filter(rid="123456789").exists())
        heartbeat = self._post_json(
            "/api/heartbeat",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
            token=token,
        )
        self.assertEqual(heartbeat.status_code, 200, heartbeat.content)

    def test_address_book_peers_match_flutter_client_contract(self):
        token = self._login("alice", "alice-pass")
        profile = self._personal_profile(self.user)
        peer = RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
            username="mira",
            hostname="studio-mac",
            platform="Mac OS",
            alias="Design workstation",
            note="Primary workstation",
            device_group_name="Design",
            login_name="mira@example.test",
            same_server=True,
            rhash="personal-hash",
        )
        tags = [
            RemoteTag.objects.create(
                profile=profile,
                tag_name=name,
            )
            for name in ("studio", "trusted")
        ]
        peer.tags.set(tags)

        response = self._post_json(
            "/api/ab/peers?current=1&pageSize=100",
            {},
            token=token,
        )

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["total"], 1)
        self.assertEqual(
            response.json()["data"][0],
            {
                "id": "765432100",
                "username": "mira",
                "hostname": "studio-mac",
                "platform": "Mac OS",
                "alias": "Design workstation",
                "tags": ["studio", "trusted"],
                "note": "Primary workstation",
                "device_group_name": "Design",
                "loginName": "mira@example.test",
                "same_server": True,
                "hash": "personal-hash",
                "password": "",
            },
        )

    def test_address_book_tag_relations_survive_merge_and_reject_cross_book_links(self):
        token = self._login("alice", "alice-pass")
        profile = self._personal_profile(self.user)
        peer = RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
        )
        old_tag = RemoteTag.objects.create(
            profile=profile,
            tag_name="old",
        )
        target_tag = RemoteTag.objects.create(
            profile=profile,
            tag_name="target",
        )
        peer.tags.add(old_tag)

        renamed = self._put_json(
            f"/api/ab/tag/rename/{profile.guid}",
            {"old": "old", "new": "target"},
            token=token,
        )

        self.assertEqual(renamed.status_code, 200, renamed.content)
        self.assertFalse(RemoteTag.objects.filter(profile=profile, tag_name="old").exists())
        self.assertEqual(
            list(peer.tags.values_list("tag_name", flat=True)),
            ["target"],
        )

        other_profile = AddressBookProfile.objects.create(
            owner=self.user,
            guid="separate-address-book",
            name="Separate",
            rule=3,
        )
        foreign_tag = RemoteTag.objects.create(
            profile=other_profile,
            tag_name="foreign",
        )
        with self.assertRaises(ValidationError):
            with transaction.atomic():
                peer.tags.add(foreign_tag)

        deleted = self._delete_json(
            f"/api/ab/tag/{profile.guid}",
            ["target"],
            token=token,
        )
        self.assertEqual(deleted.status_code, 200, deleted.content)
        self.assertFalse(peer.tags.exists())
        self.assertFalse(RemoteTag.objects.filter(pk=target_tag.pk).exists())

    def test_address_book_credentials_are_authenticated_encrypted_at_rest(self):
        profile = self._personal_profile(self.user)
        first = RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
            rhash="same-sensitive-value",
            password="same-sensitive-value",
        )
        second = RemotePeer.objects.create(
            profile=profile,
            rid="765432101",
            rhash="same-sensitive-value",
        )
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT rhash, password FROM api_remotepeer WHERE id = %s",
                [first.pk],
            )
            raw_rhash, raw_password = cursor.fetchone()
            cursor.execute(
                "SELECT rhash FROM api_remotepeer WHERE id = %s",
                [second.pk],
            )
            second_raw_rhash = cursor.fetchone()[0]
        envelope_prefix = f"{FIELD_PREFIX}{project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID}:"
        self.assertTrue(raw_rhash.startswith(envelope_prefix))
        self.assertTrue(raw_password.startswith(envelope_prefix))
        self.assertNotIn("same-sensitive-value", raw_rhash)
        self.assertNotEqual(raw_rhash, raw_password)
        self.assertNotEqual(raw_rhash, second_raw_rhash)

        first.refresh_from_db()
        self.assertEqual(first.rhash, "same-sensitive-value")
        self.assertEqual(first.password, "same-sensitive-value")

        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE api_remotepeer SET rhash = %s WHERE id = %s",
                [f"{envelope_prefix}AAAA", first.pk],
            )
        with self.assertRaises(ValidationError):
            first.refresh_from_db()

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_admin_forms_and_exports_do_not_render_connection_credentials(self):
        profile = AddressBookProfile.objects.create(
            owner=self.user,
            guid="team-operations",
            name="Operations",
            rule=3,
        )
        peer_secret = "peer-secret-that-must-not-render"
        peer = RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
            password=peer_secret,
        )
        device_secret = "device-secret-that-must-not-render"
        device = self._device(
            owner=self.user,
            rid="765432101",
            uuid=device_uuid("admin-secret-device"),
            address_book_password=device_secret,
        )
        self.client.force_login(self.admin)

        for path, secret in (
            (f"/admin/api/remotepeer/{peer.pk}/change/", peer_secret),
            (f"/admin/api/remotedevice/{device.pk}/change/", device_secret),
        ):
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200, response.content)
                self.assertNotContains(response, secret)
                self.assertContains(response, 'type="password"')

        self.client.force_login(self.user)
        exported = self.client.get(
            f"/api/ab_book_export?guid={profile.guid}&kind=peers&format=csv",
        )
        self.assertEqual(exported.status_code, 200, exported.content)
        self.assertNotIn(peer_secret.encode(), exported.content)

    def test_device_inventory_export_uses_a_safe_versioned_allowlist(self):
        credential_canary = "DEVICE-EXPORT-CREDENTIAL-CANARY"
        uuid_canary = device_uuid("device-export-uuid-canary")
        public_key_hash_canary = hashlib.sha256(b"device-export-public-key-canary").hexdigest()
        device = self._device(
            owner=self.user,
            rid="765432188",
            uuid=uuid_canary,
            public_key_hash=public_key_hash_canary,
            address_book_password=credential_canary,
            version="DEVICE-EXPORT-VERSION-CANARY",
            os="DEVICE-EXPORT-OS-CANARY",
        )

        self.client.force_login(self.user)
        denied = self.client.get("/api/down_peers")
        self.assertEqual(denied.status_code, 302)
        self.assertEqual(denied.headers["Location"], "/api/work")

        self.client.force_login(self.admin)
        with (
            CaptureQueriesContext(connection) as export_queries,
            self.assertLogs("api.views_front", level="INFO") as export_logs,
        ):
            response = self.client.get("/api/down_peers")

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.streaming)
        response_body = b"".join(response.streaming_content)
        for closer in response._resource_closers:
            closer()
        response._resource_closers.clear()
        workbook = load_workbook(BytesIO(response_body), read_only=True, data_only=False)
        try:
            worksheet = workbook.active
            exported_rows = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        self.assertEqual(len(exported_rows), 2)
        headers = tuple(exported_rows[0])
        values = tuple(exported_rows[1])
        for forbidden_header in (
            "id",
            "uuid",
            "public_key_hash",
            "deployment_generation",
            "policy_generation",
            "address_book_password",
        ):
            with self.subTest(header=forbidden_header):
                self.assertNotIn(forbidden_header, headers)
        for canary in (credential_canary, uuid_canary, public_key_hash_canary):
            with self.subTest(canary=canary):
                self.assertNotIn(canary, values)
        export_sql = "\n".join(query["sql"] for query in export_queries.captured_queries)
        for forbidden_column in ("address_book_password", "uuid", "public_key_hash"):
            with self.subTest(sql_column=forbidden_column):
                self.assertNotIn(f'"api_remotedevice"."{forbidden_column}"', export_sql)

        expected_headers = (
            "rid",
            "owner_name",
            "device_group_name",
            "strategy_name",
            "version",
            "os",
            "enabled",
            "status",
            "update_time",
        )
        self.assertEqual(headers, expected_headers)
        exported = dict(zip(headers, values, strict=True))
        self.assertEqual(exported["rid"], device.rid)
        self.assertEqual(exported["owner_name"], self.user.username)
        self.assertEqual(exported["version"], "DEVICE-EXPORT-VERSION-CANARY")
        self.assertEqual(exported["os"], "DEVICE-EXPORT-OS-CANARY")
        self.assertIs(exported["enabled"], True)
        self.assertEqual(response.headers.get("Cache-Control"), "no-store, private")
        self.assertEqual(response.headers.get("Pragma"), "no-cache")
        self.assertEqual(response.headers.get("X-Camellia-Export-Schema"), "device-inventory-v1")
        self.assertEqual(response.headers.get("Content-Disposition"), 'attachment; filename="DeviceInfo-v1.xlsx"')

        log_output = "\n".join(export_logs.output)
        export_payload = next(
            payload
            for payload in (json.loads(record.getMessage()) for record in export_logs.records)
            if payload.get("event") == "front_export_xlsx"
        )
        self.assertEqual(export_payload["attributes"]["schema"], "device-inventory-v1")
        self.assertEqual(export_payload["attributes"]["count"], 1)
        for canary in (credential_canary, uuid_canary, public_key_hash_canary):
            with self.subTest(log_canary=canary):
                self.assertNotIn(canary, log_output)

    def test_expired_authentication_state_cleanup_is_dry_run_safe_and_idempotent(self):
        recording_root = self.enterContext(tempfile.TemporaryDirectory())
        self.enterContext(
            override_settings(
                RECORD_UPLOAD_ROOT=Path(recording_root),
                RECORD_UPLOAD_REQUIRE_MOUNT=False,
                RECORD_UPLOAD_VOLUME_RESERVE_BYTES=0,
                RECORD_UPLOAD_VOLUME_RESERVE_INODES=0,
                RECORD_UPLOAD_CAPABILITY_CACHE_SECONDS=0,
            )
        )
        now = timezone.now()
        old = now - datetime.timedelta(days=60)
        device = self._device(owner=self.user)
        token = RemoteToken.objects.create(
            device=device,
            subject_user=self.user,
            access_token="a" * 64,
            expires_at=now - datetime.timedelta(minutes=1),
        )
        attempt = LoginAttempt.objects.create(ip="192.0.2.1", username="alice")
        LoginAttempt.objects.filter(pk=attempt.pk).update(
            created_at=now - datetime.timedelta(hours=1),
        )
        admission_lock = LoginAdmissionLock.objects.create(ip="192.0.2.1")
        LoginAdmissionLock.objects.filter(pk=admission_lock.pk).update(
            updated_at=now - datetime.timedelta(hours=1),
        )
        oidc = OidcPendingAuth.objects.create(
            state="expired-state",
            poll_code_hash="b" * 64,
            provider="example",
            request_ip="192.0.2.1",
            nonce="expired-nonce",
            code_verifier="expired-verifier",
        )
        OidcPendingAuth.objects.filter(pk=oidc.pk).update(created_at=old)
        recent_expired_share = ShareLink.objects.create(
            creator=self.user,
            shash="c" * 64,
            token_prefix="recent",
            expires_at=now - datetime.timedelta(minutes=1),
        )
        retained_share = ShareLink.objects.create(
            creator=self.user,
            shash="d" * 64,
            token_prefix="retained",
            expires_at=now + datetime.timedelta(days=1),
            is_used=True,
            used_at=old,
            used_by=self.user,
        )
        ShareLink.objects.filter(pk=retained_share.pk).update(create_time=old)
        rate_bucket = RequestRateBucket.objects.create(
            key_hash="e" * 64,
            scope="source",
            group="read",
            window_seconds=60,
            used=1,
            expires_at=now - datetime.timedelta(seconds=1),
        )
        rate_lease = RequestRateLease.objects.create(
            request_id="f" * 32,
            key_hash="f" * 64,
            scope="source",
            group="read",
            expires_at=now - datetime.timedelta(seconds=1),
        )

        output = StringIO()
        call_command("purge_expired_state", "--dry-run", stdout=output)
        self.assertTrue(RemoteToken.objects.filter(pk=token.pk).exists())
        self.assertTrue(LoginAttempt.objects.filter(pk=attempt.pk).exists())
        self.assertTrue(LoginAdmissionLock.objects.filter(pk=admission_lock.pk).exists())
        self.assertTrue(OidcPendingAuth.objects.filter(pk=oidc.pk).exists())
        self.assertFalse(ShareLink.objects.get(pk=recent_expired_share.pk).is_expired)
        self.assertTrue(ShareLink.objects.filter(pk=retained_share.pk).exists())
        self.assertTrue(RequestRateBucket.objects.filter(pk=rate_bucket.pk).exists())
        self.assertTrue(RequestRateLease.objects.filter(pk=rate_lease.pk).exists())
        self.assertTrue(json.loads(output.getvalue())["dry_run"])

        call_command("purge_expired_state", stdout=StringIO())
        self.assertFalse(RemoteToken.objects.filter(pk=token.pk).exists())
        self.assertFalse(LoginAttempt.objects.filter(pk=attempt.pk).exists())
        self.assertFalse(LoginAdmissionLock.objects.filter(pk=admission_lock.pk).exists())
        self.assertFalse(OidcPendingAuth.objects.filter(pk=oidc.pk).exists())
        recent_expired_share.refresh_from_db()
        self.assertTrue(recent_expired_share.is_expired)
        self.assertFalse(ShareLink.objects.filter(pk=retained_share.pk).exists())
        self.assertFalse(RequestRateBucket.objects.filter(pk=rate_bucket.pk).exists())
        self.assertFalse(RequestRateLease.objects.filter(pk=rate_lease.pk).exists())

        second = StringIO()
        call_command("purge_expired_state", stdout=second)
        result = json.loads(second.getvalue())
        self.assertEqual(
            {key: value for key, value in result.items() if key != "dry_run"},
            {
                "expired_access_tokens": 0,
                "expired_device_proof_challenges": 0,
                "expired_device_recovery_approvals": 0,
                "expired_oidc_sessions": 0,
                "expired_share_links_marked": 0,
                "login_attempts": 0,
                "login_admission_locks": 0,
                "retained_share_links": 0,
                "request_rate_buckets": 0,
                "request_rate_leases": 0,
                "expired_management_batch_operations": 0,
                "management_batch_operations_purged": 0,
                "management_batch_operations_remaining": 0,
                "recording_active_expired": 0,
                "recording_finalized_purged": 0,
                "recording_aborted_purged": 0,
                "recording_orphans_quarantined": 0,
                "recording_quarantine_purged": 0,
                "audit_connections_purged": 0,
                "audit_connections_expired": 0,
                "legacy_file_audits_purged": 0,
                "legacy_alarm_audits_purged": 0,
            },
        )

    def test_admin_can_manage_users_and_disabled_users_cannot_login(self):
        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )

        created = self._post_json(
            "/api/users",
            {
                "name": "bob",
                "password": "Bob-API-test-4e!9",
                "email": "bob@example.test",
                "group_name": "ops",
            },
            token=admin_token,
        )
        self.assertEqual(created.status_code, 200, created.content)
        user_guid = created.json()["guid"]

        duplicate = self._post_json(
            "/api/users",
            {"name": "bob", "password": "Bob-API-test-4e!9"},
            token=admin_token,
        )
        self.assertEqual(duplicate.status_code, 409)

        listed = self.client.get("/api/users?name=bob", **self._auth_headers(admin_token))
        self.assertEqual(listed.status_code, 200, listed.content)
        self.assertEqual(listed.json()["total"], 1)
        self.assertEqual(listed.json()["data"][0]["group_name"], "ops")
        self.assertEqual(listed.json()["data"][0]["group_names"], ["ops"])
        self.assertEqual(
            list(UserProfile.objects.get(username="bob").groups.values_list("name", flat=True)),
            ["ops"],
        )

        disabled = self._post_json(f"/api/users/{user_guid}/disable", {}, token=admin_token)
        self.assertEqual(disabled.status_code, 200, disabled.content)
        denied = self._post_json(
            "/api/login",
            {
                "username": "bob",
                "password": "Bob-API-test-4e!9",
                "id": "800000001",
                "uuid": device_uuid("bob-device"),
            },
        )
        # Inactive and unknown accounts deliberately share the same response so
        # the login endpoint does not become an account-status oracle.
        self.assertEqual(denied.status_code, 401)

        enabled = self._post_json(f"/api/users/{user_guid}/enable", {}, token=admin_token)
        self.assertEqual(enabled.status_code, 200, enabled.content)
        self.assertTrue(enabled.json()["status"])

    def test_real_user_group_membership_grants_address_book_rule_access(self):
        group = Group.objects.create(name="design")
        recipient = UserProfile.objects.create_user(
            username="bob",
            password="bob-pass",
        )
        recipient.groups.add(group)
        profile = AddressBookProfile.objects.create(
            owner=self.user,
            guid="team-design",
            name="Design team",
            rule=3,
        )
        AddressBookRule.objects.create(
            profile=profile,
            group=group,
            rule=1,
        )
        RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
            alias="Design workstation",
        )
        token = self._login(
            "bob",
            "bob-pass",
            rid="800000001",
            uuid=device_uuid("bob-device"),
        )

        profiles = self._post_json(
            "/api/ab/shared/profiles",
            {},
            token=token,
        )
        peers = self._post_json(
            f"/api/ab/peers?ab={profile.guid}",
            {},
            token=token,
        )

        self.assertEqual(profiles.status_code, 200, profiles.content)
        self.assertEqual(
            profiles.json()["data"],
            [
                {
                    "guid": "team-design",
                    "name": "Design team",
                    "owner": "alice",
                    "note": "",
                    "info": {},
                    "rule": 1,
                }
            ],
        )
        self.assertEqual(peers.status_code, 200, peers.content)
        self.assertEqual(peers.json()["data"][0]["id"], "765432100")

    def test_accessible_inventory_never_joins_an_unowned_device_by_id(self):
        foreign_owner = UserProfile.objects.create_user(
            username="bob",
            password="bob-pass",
        )
        foreign_group = DeviceGroup.objects.create(name="foreign-group")
        self._device(
            owner=foreign_owner,
            rid="765432100",
            uuid=device_uuid("foreign-device"),
            hostname="private-hostname",
            username="private-user",
            device_group=foreign_group,
        )
        personal_profile = self._personal_profile(self.user)
        RemotePeer.objects.create(
            profile=personal_profile,
            rid="765432100",
            hostname="address-book-hostname",
            username="address-book-user",
            device_group_name="local-label",
        )
        token = self._login("alice", "alice-pass")
        own_group = DeviceGroup.objects.create(name="own-group")
        own_device = RemoteDevice.objects.get(rid="123456789")
        own_device.device_group = own_group
        own_device.save(update_fields=["device_group"])
        RemoteDevice.objects.filter(pk=own_device.pk).update(
            update_time=timezone.now() - datetime.timedelta(days=1),
        )

        peers = self.client.get(
            "/api/peers?status=1",
            **self._auth_headers(token),
        )
        groups = self.client.get(
            "/api/device-group/accessible",
            **self._auth_headers(token),
        )

        self.assertEqual(peers.status_code, 200, peers.content)
        peer_items = {item["id"]: item for item in peers.json()["data"]}
        self.assertEqual(
            set(peer_items),
            {"123456789", "765432100"},
        )
        self.assertEqual(
            peer_items["765432100"]["info"],
            {
                "username": "address-book-user",
                "os": "",
                "device_name": "address-book-hostname",
            },
        )
        self.assertEqual(
            peer_items["765432100"]["device_group_name"],
            "local-label",
        )
        self.assertEqual(peer_items["765432100"]["user"], "alice")
        self.assertEqual(groups.status_code, 200, groups.content)
        self.assertEqual(
            [item["name"] for item in groups.json()["data"]],
            ["own-group"],
        )

    def test_device_groups_and_strategies_drive_heartbeat_contract(self):
        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )
        self._device(owner=self.user)
        device_token = self._login("alice", "alice-pass")

        created_group = self._post_json(
            "/api/device-groups",
            {
                "name": "ops",
                "note": "Operations devices",
                "allowed_incomings": ["admin"],
            },
            token=admin_token,
        )
        self.assertEqual(created_group.status_code, 200, created_group.content)
        group_guid = created_group.json()["guid"]

        assigned_devices = self._post_json(
            f"/api/device-groups/{group_guid}",
            ["123456789"],
            token=admin_token,
        )
        self.assertEqual(assigned_devices.status_code, 200, assigned_devices.content)
        self.assertEqual(assigned_devices.json()["updated"], 1)

        strategy = StrategyProfile.objects.create(
            name="high-quality",
            config_options={"quality": "best", "enable-audio": "Y"},
        )
        assigned_strategy = self._post_json(
            "/api/strategies/assign",
            {"strategy": str(strategy.guid), "groups": [group_guid]},
            token=admin_token,
        )
        self.assertEqual(assigned_strategy.status_code, 200, assigned_strategy.content)
        self.assertEqual(assigned_strategy.json()["groups"], 1)

        heartbeat = self._post_json(
            "/api/heartbeat",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID, "modified_at": 0},
            token=device_token,
        )
        self.assertEqual(heartbeat.status_code, 200, heartbeat.content)
        initial_policy = heartbeat.json()["managed_policy"]
        self.assertEqual(initial_policy["version"], 1)
        self.assertEqual(initial_policy["id"], "123456789")
        self.assertEqual(initial_policy["uuid"], DEFAULT_DEVICE_UUID)
        self.assertEqual(initial_policy["config_options"]["quality"], "best")
        canonical = json.dumps(
            initial_policy["config_options"],
            sort_keys=True,
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode()
        self.assertEqual(initial_policy["digest"], hashlib.sha256(canonical).hexdigest())
        self.assertNotIn("strategy", heartbeat.json())
        self.assertNotIn("modified_at", heartbeat.json())

        disabled = self._put_json(
            f"/api/strategies/{strategy.guid}/status",
            {"enabled": False},
            token=admin_token,
        )
        self.assertEqual(disabled.status_code, 200, disabled.content)
        self.assertFalse(disabled.json()["enabled"])

        heartbeat = self._post_json(
            "/api/heartbeat",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID, "modified_at": 0},
            token=device_token,
        )
        self.assertEqual(heartbeat.status_code, 200, heartbeat.content)
        tombstone = heartbeat.json()["managed_policy"]
        self.assertGreater(tombstone["generation"], initial_policy["generation"])
        self.assertEqual(tombstone["config_options"], {})

    def test_strategy_precedence_and_device_policy_is_admin_only(self):
        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )
        device = self._device(owner=self.user)
        device_token = self._login("alice", "alice-pass")
        user_strategy = StrategyProfile.objects.create(
            name="user-policy",
            config_options={"source": "user"},
        )
        group_strategy = StrategyProfile.objects.create(
            name="group-policy",
            config_options={"source": "group"},
        )
        direct_strategy = StrategyProfile.objects.create(
            name="direct-policy",
            config_options={"source": "device"},
        )
        group = DeviceGroup.objects.create(
            name="ops",
            strategy=group_strategy,
        )
        device.device_group = group
        device.strategy = direct_strategy
        device.save(update_fields=["device_group", "strategy"])
        self.user.strategy = user_strategy
        self.user.save(update_fields=["strategy"])

        heartbeat = self._post_json(
            "/api/heartbeat",
            {
                "id": device.rid,
                "uuid": device.uuid,
                "modified_at": 0,
            },
            token=device_token,
        )
        self.assertEqual(
            heartbeat.json()["managed_policy"]["config_options"]["source"],
            "device",
        )

        cleared = self._post_json(
            f"/api/devices/{device.pk}/assign",
            {"type": "strategy_name", "value": ""},
            token=admin_token,
        )
        self.assertEqual(cleared.status_code, 200, cleared.content)
        heartbeat = self._post_json(
            "/api/heartbeat",
            {
                "id": device.rid,
                "uuid": device.uuid,
                "modified_at": 0,
            },
            token=device_token,
        )
        self.assertEqual(
            heartbeat.json()["managed_policy"]["config_options"]["source"],
            "group",
        )

        group.strategy = None
        group.save(update_fields=["strategy"])
        heartbeat = self._post_json(
            "/api/heartbeat",
            {
                "id": device.rid,
                "uuid": device.uuid,
                "modified_at": 0,
            },
            token=device_token,
        )
        self.assertEqual(
            heartbeat.json()["managed_policy"]["config_options"]["source"],
            "user",
        )

        ignored = self._post_json(
            "/api/sysinfo",
            {
                "id": device.rid,
                "uuid": device.uuid,
                "strategy_name": direct_strategy.name,
                "device_group_name": "other",
            },
            token=device_token,
        )
        self.assertEqual(ignored.status_code, 200, ignored.content)
        device.refresh_from_db()
        self.assertIsNone(device.strategy_id)
        self.assertEqual(device.device_group_id, group.id)

    def test_heartbeat_fails_closed_for_corrupt_strategy_data(self):
        device = self._device(owner=self.user)
        device_token = self._login("alice", "alice-pass")
        strategy = StrategyProfile.objects.create(
            name="strict-policy",
            config_options={"source": "valid"},
        )
        device.strategy = strategy
        device.save(update_fields=["strategy"])
        StrategyProfile.objects.filter(pk=strategy.pk).update(
            config_options={"unexpected": ["non-string-value"]},
        )

        heartbeat = self._post_json(
            "/api/heartbeat",
            {"id": device.rid, "uuid": device.uuid, "modified_at": 0},
            token=device_token,
        )

        self.assertEqual(heartbeat.status_code, 503, heartbeat.content)
        self.assertEqual(heartbeat.json(), {"error": "Invalid strategy configuration"})

    def test_disabled_devices_are_blocked_by_login_and_heartbeat(self):
        device = self._device(owner=self.user)
        token = self._login("alice", "alice-pass")
        device.is_active = False
        device.save(update_fields=["is_active"])

        login = self._post_json(
            "/api/login",
            {
                "username": "alice",
                "password": "alice-pass",
                "id": "123456789",
                "uuid": DEFAULT_DEVICE_UUID,
            },
        )
        self.assertEqual(login.status_code, 403)

        heartbeat = self._post_json(
            "/api/heartbeat",
            {"id": "123456789", "uuid": DEFAULT_DEVICE_UUID},
            token=token,
        )
        # Disabling the device invalidates its bearer before endpoint-specific
        # authorization and does not disclose device state to the old token.
        self.assertEqual(heartbeat.status_code, 401)

    @override_settings(DEVICE_VERIFICATION_TOKEN="v" * 48)
    def test_device_deployment_is_bound_to_uuid_key_and_active_owner(self):
        managed_uuid = base64.b64encode(b"managed-device").decode()
        token = self._login("alice", "alice-pass", uuid=managed_uuid)
        signing_key = SigningKey.generate()
        public_key = base64.b64encode(bytes(signing_key.verify_key)).decode()
        proof = self._device_proof("deploy", signing_key, "123456789", managed_uuid, token=token)
        deployed = self._post_json(
            "/api/devices/deploy",
            {
                "id": "123456789",
                "uuid": managed_uuid,
                "pk": public_key,
                "device_proof": proof,
            },
            token=token,
        )
        self.assertEqual(deployed.status_code, 200, deployed.content)
        device = RemoteDevice.objects.get(rid="123456789")
        self.assertEqual(device.uuid, managed_uuid)
        self.assertEqual(device.public_key_hash, hashlib.sha256(bytes(signing_key.verify_key)).hexdigest())

        payload = {
            "id": device.rid,
            "uuid": device.uuid,
            "public_key_hash": device.public_key_hash,
            "request_nonce": base64.b64encode(b"v" * 32).decode("ascii"),
        }
        self.assertEqual(self._post_json("/api/devices/verify-deployment", payload).status_code, 401)
        verified = self._post_json(
            "/api/devices/verify-deployment",
            payload,
            token="v" * 48,
        )
        self.assertEqual(verified.status_code, 200, verified.content)
        self.assertEqual(verified.json()["deployment_generation"], 1)

        payload["public_key_hash"] = "0" * 64
        denied = self._post_json(
            "/api/devices/verify-deployment",
            payload,
            token="v" * 48,
        )
        self.assertEqual(denied.status_code, 404)

        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )
        disabled = self._post_json(f"/api/devices/{device.pk}/disable", {}, token=admin_token)
        self.assertEqual(disabled.status_code, 200, disabled.content)
        payload["public_key_hash"] = device.public_key_hash
        revoked = self._post_json(
            "/api/devices/verify-deployment",
            payload,
            token="v" * 48,
        )
        self.assertEqual(revoked.status_code, 404)
        self.assertFalse(RemoteToken.objects.filter(access_token=hashlib.sha256(token.encode()).hexdigest()).exists())

    @override_settings(DEVICE_VERIFICATION_TOKEN="v" * 48)
    def test_device_deployment_rejects_noncanonical_identity_and_id_takeover(self):
        first_uuid = base64.b64encode(b"first-device").decode()
        token = self._login("alice", "alice-pass", uuid=first_uuid)
        signing_key = SigningKey.generate()
        public_key = base64.b64encode(bytes(signing_key.verify_key)).decode()
        malformed = self._post_json(
            "/api/devices/deploy",
            {"id": "123456789", "uuid": "not-base64", "pk": public_key},
            token=token,
        )
        self.assertEqual(malformed.status_code, 400)

        first = self._post_json(
            "/api/devices/deploy",
            {
                "id": "123456789",
                "uuid": first_uuid,
                "pk": public_key,
                "device_proof": self._device_proof(
                    "deploy",
                    signing_key,
                    "123456789",
                    first_uuid,
                    token=token,
                ),
            },
            token=token,
        )
        self.assertEqual(first.status_code, 200, first.content)

        bob = UserProfile.objects.create_user(username="bob", password="bob-pass")
        second_uuid = base64.b64encode(b"second-device").decode()
        bob_token = self._login(
            bob.username,
            "bob-pass",
            rid="777777777",
            uuid=second_uuid,
        )
        takeover = self._post_json(
            "/api/devices/deploy",
            {
                "id": "123456789",
                "uuid": second_uuid,
                "pk": base64.b64encode(bytes(reversed(range(32)))).decode(),
            },
            token=bob_token,
        )
        self.assertEqual(takeover.status_code, 409)

    def test_administrator_cannot_impersonate_an_owned_device(self):
        device = self._device(owner=self.user)

        login = self._post_json(
            "/api/login",
            {
                "username": "admin",
                "password": "admin-pass",
                "id": device.rid,
                "uuid": device.uuid,
            },
        )
        self.assertEqual(login.status_code, 403, login.content)
        self.assertFalse(
            RemoteToken.objects.filter(
                device__owner=self.admin,
                device__rid=device.rid,
                device__uuid=device.uuid,
            ).exists()
        )

    def test_inactive_owner_invalidates_device_heartbeat(self):
        device = self._device(owner=self.user)
        token = self._login("alice", "alice-pass")
        self.user.is_active = False
        self.user.save(update_fields=["is_active"])

        response = self._post_json(
            "/api/heartbeat",
            {"id": device.rid, "uuid": device.uuid},
            token=token,
        )
        self.assertEqual(response.status_code, 401, response.content)
        self.assertFalse(RemoteToken.objects.filter(device__owner=self.user).exists())

    @override_settings(ALLOW_REGISTRATION=True)
    def test_first_public_registration_never_bootstraps_an_administrator(self):
        UserProfile.objects.all().delete()
        response = self.client.post(
            "/api/user_action?action=register",
            {
                "user": "first-user",
                "pwd": "strong-pass",
                "repassword": "strong-pass",
            },
        )
        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.json()["code"], 1)
        user = UserProfile.objects.get(username="first-user")
        self.assertFalse(user.is_admin)
        self.assertFalse(user.is_superuser)

        mismatch = self.client.post(
            "/api/user_action?action=register",
            {
                "user": "second-user",
                "pwd": "strong-pass",
                "repassword": "different-pass",
            },
        )
        self.assertEqual(mismatch.status_code, 400, mismatch.content)
        self.assertEqual(mismatch.json()["code"], 0)
        self.assertFalse(UserProfile.objects.filter(username="second-user").exists())

    @override_settings(PLUGIN_SIGNING_KEY="00" * 32)
    def test_plugin_raw_signing_oracle_is_not_routed(self):
        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )
        payloads = (
            {"msg": []},
            {"msg": [True, False, 255]},
            {"msg": [9, 8, 7], "plugin_id": "alpha", "version": "1.0.0"},
            {"msg": [9, 8, 7], "plugin_id": "beta", "version": "2.0.0"},
        )

        for payload in payloads:
            with self.subTest(payload=payload):
                response = self._post_json(
                    "/lic/web/api/plugin-sign",
                    payload,
                    token=admin_token,
                )
                self.assertEqual(response.status_code, 404, response.content)

    def test_device_group_delete_detaches_devices(self):
        admin_token = self._login(
            "admin",
            "admin-pass",
            rid="900000001",
            uuid=device_uuid("admin-device"),
        )
        group = DeviceGroup.objects.create(name="ops")
        device = self._device(owner=self.user, device_group=group)

        response = self._delete_json(f"/api/device-groups/{group.guid}", token=admin_token)
        self.assertEqual(response.status_code, 200, response.content)
        device.refresh_from_db()
        self.assertIsNone(device.device_group_id)

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_front_device_page_uses_modern_device_owner_relation(self):
        self._device(
            owner=self.user,
            rid="765432100",
            uuid=device_uuid("owned-device"),
        )
        self.client.force_login(self.user)

        response = self.client.get("/api/work")

        self.assertEqual(response.status_code, 200, response.content)
        self.assertEqual(response.context["page_obj"].paginator.count, 1)
        self.assertContains(response, "765432100")

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_front_device_pages_merge_address_book_metadata_and_expose_unknown_state(self):
        self._device(
            owner=self.user,
            rid="765432100",
            uuid=device_uuid("owned-device"),
        )
        personal_profile = self._personal_profile(self.user)
        work_tag = RemoteTag.objects.create(
            profile=personal_profile,
            tag_name="work",
        )
        mobile_tag = RemoteTag.objects.create(
            profile=personal_profile,
            tag_name="mobile",
        )
        primary_peer = RemotePeer.objects.create(
            profile=personal_profile,
            rid="765432100",
            username="desktop-user",
            hostname="desktop",
            alias="Primary desktop",
            platform="Linux",
            rhash="secret-hash",
        )
        address_book_peer = RemotePeer.objects.create(
            profile=personal_profile,
            rid="765432101",
            username="address-book-user",
            hostname="unknown",
            alias="Address book only",
            platform="Android",
            rhash="",
        )
        primary_peer.tags.add(work_tag)
        address_book_peer.tags.add(mobile_tag)
        self.client.force_login(self.user)

        work_response = self.client.get("/api/work")
        home_response = self.client.get("/api/home")

        self.assertEqual(work_response.status_code, 200, work_response.content)
        items = {item["rid"]: item for item in work_response.context["page_obj"]}
        self.assertEqual(set(items), {"765432100", "765432101"})
        self.assertEqual(items["765432100"]["alias"], "Primary desktop")
        self.assertEqual(items["765432100"]["platform"], "Linux")
        self.assertEqual(items["765432101"]["status"], "未知状态")
        self.assertEqual(home_response.context["summary"], {"total": 2, "online": 1, "offline": 0, "unknown": 1})

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_front_device_pages_do_not_expose_another_owners_inventory_by_rid(self):
        self._device(
            owner=self.user,
            rid="765432100",
            uuid=device_uuid("alice-owned-device"),
            cpu="ALICE-CPU-CANARY",
            hostname="ALICE-HOST-CANARY",
            memory="ALICE-MEMORY-CANARY",
            version="ALICE-VERSION-CANARY",
            ip_address="192.0.2.10",
        )
        bob = UserProfile.objects.create_user(username="bob-inventory-owner", password="bob-pass")
        bob_device = self._device(
            owner=bob,
            rid="765432199",
            uuid=device_uuid("bob-owned-device"),
            cpu="BOB-CPU-INVENTORY-CANARY",
            hostname="BOB-HOST-INVENTORY-CANARY",
            memory="BOB-MEMORY-INVENTORY-CANARY",
            os="BOB-OS-INVENTORY-CANARY",
            username="BOB-USER-INVENTORY-CANARY",
            version="BOB-VERSION-INVENTORY-CANARY",
            ip_address="203.0.113.77",
        )
        bob_update_time = datetime.datetime(2037, 2, 3, 4, 5, tzinfo=datetime.UTC)
        RemoteDevice.objects.filter(pk=bob_device.pk).update(update_time=bob_update_time)
        bob_device.refresh_from_db()

        personal_profile = self._personal_profile(self.user)
        RemotePeer.objects.create(
            profile=personal_profile,
            rid=bob_device.rid,
            username="ALICE-SAVED-USER",
            hostname="ALICE-SAVED-HOST",
            alias="Alice saved Bob",
            note="Alice private note",
            platform="ALICE-SAVED-PLATFORM",
        )

        self.client.force_login(self.user)
        with self.assertLogs("api.views_front", level="INFO") as alice_logs:
            alice_work = self.client.get("/api/work")
            alice_home = self.client.get("/api/home")
        self.client.force_login(bob)
        bob_work = self.client.get("/api/work")
        self.client.force_login(self.admin)
        admin_work = self.client.get("/api/work?show_type=admin")

        for response in (alice_work, alice_home, bob_work, admin_work):
            self.assertEqual(response.status_code, 200, response.content)

        bob_owned_item = {item["rid"]: item for item in bob_work.context["page_obj"]}[bob_device.rid]
        self.assertEqual(bob_owned_item["cpu"], "BOB-CPU-INVENTORY-CANARY")
        self.assertEqual(bob_owned_item["hostname"], "BOB-HOST-INVENTORY-CANARY")
        admin_item = {item["rid"]: item for item in admin_work.context["page_obj"]}[bob_device.rid]
        self.assertEqual(admin_item["cpu"], "BOB-CPU-INVENTORY-CANARY")
        self.assertEqual(admin_item["ip_address"], "203.0.113.77")

        alice_items = {item["rid"]: item for item in alice_work.context["page_obj"]}
        self.assertEqual(set(alice_items), {"765432100", bob_device.rid})
        self.assertEqual(alice_items["765432100"]["cpu"], "ALICE-CPU-CANARY")
        address_book_item = alice_items[bob_device.rid]
        self.assertEqual(address_book_item["alias"], "Alice saved Bob")
        self.assertEqual(address_book_item["note"], "Alice private note")
        self.assertEqual(address_book_item["platform"], "ALICE-SAVED-PLATFORM")
        self.assertEqual(address_book_item["username"], "ALICE-SAVED-USER")
        self.assertEqual(address_book_item["hostname"], "ALICE-SAVED-HOST")
        self.assertEqual(address_book_item["status"], "未知状态")
        for field in ("version", "os", "cpu", "memory", "ip_address", "create_time", "update_time"):
            with self.subTest(field=field):
                self.assertEqual(address_book_item[field], "")

        inventory_canaries = (
            "BOB-CPU-INVENTORY-CANARY",
            "BOB-HOST-INVENTORY-CANARY",
            "BOB-MEMORY-INVENTORY-CANARY",
            "BOB-OS-INVENTORY-CANARY",
            "BOB-USER-INVENTORY-CANARY",
            "BOB-VERSION-INVENTORY-CANARY",
            "203.0.113.77",
            "2037-02-03 04:05",
        )
        for response in (alice_work, alice_home):
            for canary in inventory_canaries:
                with self.subTest(path=response.request["PATH_INFO"], canary=canary):
                    self.assertNotContains(response, canary)
        alice_log_output = "\n".join(alice_logs.output)
        for canary in inventory_canaries:
            with self.subTest(channel="log", canary=canary):
                self.assertNotIn(canary, alice_log_output)
        self.assertContains(alice_work, "Alice saved Bob")
        self.assertContains(alice_work, "Alice private note")
        self.assertContains(alice_work, "ALICE-SAVED-PLATFORM")
        self.assertContains(alice_work, "未知状态")
        self.assertContains(bob_work, "BOB-CPU-INVENTORY-CANARY")
        self.assertContains(admin_work, "BOB-CPU-INVENTORY-CANARY")

        bob_device.owner = self.user
        bob_device.save(update_fields=("owner",))
        self.client.force_login(self.user)
        transferred_work = self.client.get("/api/work")
        self.assertContains(transferred_work, "BOB-CPU-INVENTORY-CANARY")
        transferred_item = {item["rid"]: item for item in transferred_work.context["page_obj"]}[bob_device.rid]
        self.assertEqual(transferred_item["status"], "在线")

        bob_device.owner = bob
        bob_device.save(update_fields=("owner",))
        revoked_work = self.client.get("/api/work")
        self.assertNotContains(revoked_work, "BOB-CPU-INVENTORY-CANARY")
        revoked_item = {item["rid"]: item for item in revoked_work.context["page_obj"]}[bob_device.rid]
        self.assertEqual(revoked_item["cpu"], "")
        self.assertEqual(revoked_item["status"], "未知状态")

    @override_settings(
        STORAGES={
            "default": {"BACKEND": "django.core.files.storage.FileSystemStorage"},
            "staticfiles": {"BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage"},
        }
    )
    def test_share_link_is_hashed_preview_only_and_single_use(self):
        profile = self._personal_profile(self.user)
        peer = RemotePeer.objects.create(
            profile=profile,
            rid="765432100",
            alias="Studio",
            rhash="shared-credential",
        )
        source_tag = RemoteTag.objects.create(
            profile=profile,
            tag_name="studio",
        )
        peer.tags.add(source_tag)
        self.client.force_login(self.user)
        created = self.client.post(
            "/api/share",
            {"data": json.dumps([{"value": str(peer.pk), "title": "ignored"}])},
        )
        self.assertEqual(created.status_code, 200, created.content)
        raw_token = created.json()["token"]
        link = ShareLink.objects.get()
        self.assertNotEqual(link.shash, raw_token)
        self.assertEqual(link.shash, hashlib.sha256(raw_token.encode()).hexdigest())
        self.assertEqual(
            list(link.peers.values_list("pk", flat=True)),
            [peer.pk],
        )

        recipient = UserProfile.objects.create_user(username="bob", password="bob-pass")
        self.client.force_login(recipient)
        preview = self.client.get(f"/api/share/{raw_token}")
        self.assertEqual(preview.status_code, 200, preview.content)
        link.refresh_from_db()
        self.assertFalse(link.is_used)
        self.assertFalse(
            RemotePeer.objects.filter(
                profile__owner=recipient,
                rid=peer.rid,
            ).exists()
        )

        accepted = self.client.post(f"/api/share/{raw_token}")
        self.assertEqual(accepted.status_code, 200, accepted.content)
        link.refresh_from_db()
        self.assertTrue(link.is_used)
        self.assertEqual(link.used_by, recipient)
        imported = RemotePeer.objects.get(
            profile__owner=recipient,
            rid=peer.rid,
        )
        self.assertEqual(imported.rhash, "shared-credential")
        imported_tag = imported.tags.get()
        self.assertEqual(imported_tag.tag_name, "studio")
        self.assertEqual(imported_tag.profile.owner, recipient)

        replay = self.client.post(f"/api/share/{raw_token}")
        self.assertEqual(replay.status_code, 404)


class SessionIntegrityTests(ApiTestMixin, TestCase):
    """Regressions for defects that let a session outlive its owner's intent."""

    def test_logout_revokes_the_device_that_asked(self):
        desktop_uuid = device_uuid("desktop")
        phone_uuid = device_uuid("phone")
        desktop = self._login("alice", "alice-pass", rid="111111111", uuid=desktop_uuid)
        phone = self._login("alice", "alice-pass", rid="222222222", uuid=phone_uuid)
        self.assertEqual(
            RemoteToken.objects.filter(device__owner=self.user).count(),
            2,
        )

        logout = self._post_json(
            "/api/logout",
            {"id": "111111111", "uuid": desktop_uuid},
            token=desktop,
        )
        self.assertEqual(logout.status_code, 200, logout.content)

        # The desktop is signed out ...
        self.assertEqual(self._post_json("/api/currentUser", {}, token=desktop).status_code, 401)
        # ... and the phone, which never asked, is still signed in.
        self.assertEqual(self._post_json("/api/currentUser", {}, token=phone).status_code, 200)

    def test_logout_cannot_revoke_a_device_by_public_identifiers(self):
        desktop_uuid = device_uuid("desktop")
        token = self._login("alice", "alice-pass", rid="111111111", uuid=desktop_uuid)
        denied = self._post_json(
            "/api/logout",
            {"id": "111111111", "uuid": desktop_uuid},
        )
        self.assertEqual(denied.status_code, 401)
        self.assertEqual(self._post_json("/api/currentUser", {}, token=token).status_code, 200)

    def test_one_token_row_per_device(self):
        desktop_uuid = device_uuid("desktop")
        first = self._login("alice", "alice-pass", rid="111111111", uuid=desktop_uuid)
        second = self._login("alice", "alice-pass", rid="111111111", uuid=desktop_uuid)

        # Re-login rotates in place instead of leaving an unrevokable second row.
        self.assertEqual(
            RemoteToken.objects.filter(
                device__owner=self.user,
                device__rid="111111111",
                device__uuid=desktop_uuid,
            ).count(),
            1,
        )
        self.assertEqual(self._post_json("/api/currentUser", {}, token=first).status_code, 401)
        self.assertEqual(self._post_json("/api/currentUser", {}, token=second).status_code, 200)

    def test_auth_body_carries_the_info_the_desktop_client_requires(self):
        response = self._post_json(
            "/api/login",
            {
                "username": "alice",
                "password": "alice-pass",
                "id": "123456789",
                "uuid": DEFAULT_DEVICE_UUID,
            },
        )
        user = response.json()["user"]
        # UserPayload has no default for `info`; omitting it makes the whole body
        # fail to deserialise and silently discards a successful OIDC login.
        self.assertIn("info", user)
        self.assertIn("login_device_whitelist", user["info"])

    @override_settings(TRUST_PROXY_HEADERS=False)
    def test_forwarded_headers_cannot_choose_the_rate_limit_bucket(self):
        # Every attempt claims a different forwarded address; with the header
        # untrusted they all land in the same bucket and still trip the lockout.
        for spoofed in range(10):
            self.client.post(
                "/api/login",
                data=json.dumps(
                    {
                        "username": "alice",
                        "password": "wrong",
                        "id": "123456",
                        "uuid": DEFAULT_DEVICE_UUID,
                    }
                ),
                content_type="application/json",
                HTTP_X_FORWARDED_FOR=f"198.51.100.{spoofed}",
            )
        locked = self.client.post(
            "/api/login",
            data=json.dumps(
                {
                    "username": "alice",
                    "password": "alice-pass",
                    "id": "123456",
                    "uuid": DEFAULT_DEVICE_UUID,
                }
            ),
            content_type="application/json",
            HTTP_X_FORWARDED_FOR="203.0.113.9",
        )
        self.assertEqual(locked.status_code, 429)

    def test_out_of_range_pagination_is_clamped(self):
        token = self._login("alice", "alice-pass")
        for query in ("current=0", "current=-5", "pageSize=-1", "current=0&pageSize=0"):
            response = self._post_json(
                f"/api/ab/peers?{query}",
                {},
                token=token,
            )
            self.assertEqual(response.status_code, 200, f"{query}: {response.content}")


class SensitiveIngestionTests(ApiTestMixin, TestCase):
    def _raw_record(self, query, body, token=None):
        return self.client.post(
            f"/api/record?{query}",
            data=body,
            content_type="application/octet-stream",
            CONTENT_LENGTH=str(len(body)),
            **self._auth_headers(token),
        )

    def test_record_upload_is_authenticated_sequential_and_device_isolated(self):
        self._device(owner=self.user)
        token = self._login("alice", "alice-pass")
        create_id = "11111111-1111-4111-8111-111111111111"
        with tempfile.TemporaryDirectory() as upload_root:
            with self.settings(
                RECORD_UPLOAD_ROOT=Path(upload_root),
                RECORD_UPLOAD_MAX_CHUNK_BYTES=1024,
                RECORD_UPLOAD_MAX_FILE_BYTES=2048,
                DATA_UPLOAD_MAX_MEMORY_SIZE=1024,
            ):
                self.assertEqual(
                    self._raw_record(
                        f"version=2&type=new&file=session.webm&create_id={create_id}",
                        b"",
                    ).status_code,
                    401,
                )
                created = self._raw_record(
                    f"version=2&type=new&file=session.webm&create_id={create_id}",
                    b"",
                    token,
                )
                self.assertEqual(created.status_code, 201, created.content)
                state = created.json()
                traversal = self._raw_record(
                    "version=2&type=new&file=../session.webm&create_id=22222222-2222-4222-8222-222222222222",
                    b"",
                    token,
                )
                self.assertEqual(traversal.status_code, 400)

                from api.views_api import _record_file_lock

                staging = next(Path(upload_root).glob("*/.uploads/*.part"))
                base_dir = staging.parent.parent
                with _record_file_lock(base_dir):
                    busy = self._raw_record(
                        "version=2&type=part"
                        f"&upload_id={state['upload_id']}&offset=0&revision=0&length=4"
                        f"&digest={hashlib.sha256(b'data').hexdigest()}"
                        "&chunk_id=33333333-3333-4333-8333-333333333333",
                        b"data",
                        token,
                    )
                self.assertEqual(busy.status_code, 423)

                first = self._raw_record(
                    "version=2&type=part"
                    f"&upload_id={state['upload_id']}&offset=0&revision=0&length=4"
                    f"&digest={hashlib.sha256(b'data').hexdigest()}"
                    "&chunk_id=33333333-3333-4333-8333-333333333333",
                    b"data",
                    token,
                )
                self.assertEqual(first.status_code, 200, first.content)
                conflict = self._raw_record(
                    "version=2&type=part"
                    f"&upload_id={state['upload_id']}&offset=0&revision=0&length=4"
                    f"&digest={hashlib.sha256(b'evil').hexdigest()}"
                    "&chunk_id=33333333-3333-4333-8333-333333333333",
                    b"evil",
                    token,
                )
                self.assertEqual(conflict.status_code, 409)

                self.assertNotIn(b"data", staging.read_bytes())
                self.assertGreater(staging.stat().st_size, len(b"data"))
                self.assertEqual(list(Path(upload_root).glob("*/session.webm")), [])

    def test_audit_ingestion_and_notes_are_scoped_to_authenticated_participants(self):
        host_uuid = device_uuid("host-device")
        self._device(owner=self.user, rid="111111111", uuid=host_uuid)
        host_token = self._login("alice", "alice-pass", rid="111111111", uuid=host_uuid)
        controller = UserProfile.objects.create_user(
            username="bob",
            password="bob-pass",
        )
        controller_uuid = device_uuid("controller-device")
        self._device(owner=controller, rid="222222222", uuid=controller_uuid)
        controller_token = self._login(
            "bob",
            "bob-pass",
            rid="222222222",
            uuid=controller_uuid,
        )
        outsider = UserProfile.objects.create_user(
            username="mallory",
            password="mallory-pass",
        )
        outsider_uuid = device_uuid("outsider-device")
        self._device(owner=outsider, rid="333333333", uuid=outsider_uuid)
        outsider_token = self._login(
            "mallory",
            "mallory-pass",
            rid="333333333",
            uuid=outsider_uuid,
        )
        new_event = {
            "version": 3,
            "event_id": str(uuid.uuid4()),
            "action": "new",
            "id": "111111111",
            "uuid": host_uuid,
            "conn_id": 7,
            "session_id": 99,
            "ip": "192.0.2.10",
            "type": 0,
        }

        self.assertEqual(self._post_json("/api/audit/conn", new_event).status_code, 401)
        created = self._post_json("/api/audit/conn", new_event, token=host_token)
        self.assertEqual(created.status_code, 201, created.content)
        audit_session_id = created.json()["audit_session_id"]
        replayed = self._post_json("/api/audit/conn", new_event, token=host_token)
        self.assertEqual(replayed.status_code, 200, replayed.content)
        self.assertEqual(ConnLog.objects.count(), 1)
        updated = self._post_json(
            "/api/audit/conn",
            {
                "version": 3,
                "event_id": str(uuid.uuid4()),
                "audit_session_id": audit_session_id,
                "id": "111111111",
                "uuid": host_uuid,
                "conn_id": 7,
                "session_id": 99,
                "peer": ["222222222", "bob"],
                "type": 0,
                "primary_auth": 3,
                "two_factor": 1,
            },
            token=host_token,
        )
        self.assertEqual(updated.status_code, 200, updated.content)
        connection_log = ConnLog.objects.get()
        self.assertEqual(connection_log.from_id, "222222222")
        self.assertEqual(connection_log.primary_auth, 3)
        self.assertEqual(connection_log.two_factor, 1)

        active = self.client.get(
            f"/api/audit/conn/active?version=3&event_id={uuid.uuid4()}&id=111111111&session_id=99&conn_type=0",
            **self._auth_headers(controller_token),
        )
        self.assertEqual(active.status_code, 200, active.content)
        guid = active.json()["audit_session_id"]
        self.assertTrue(guid)
        connection_log.refresh_from_db()
        self.assertEqual(str(connection_log.guid), guid)
        self.assertEqual(connection_log.actor_id, controller.id)

        direct_note = self._post_json(
            "/api/audit/conn",
            {
                "version": 3,
                "event_id": str(uuid.uuid4()),
                "audit_session_id": audit_session_id,
                "note": "during session",
            },
            token=controller_token,
        )
        self.assertEqual(direct_note.status_code, 200, direct_note.content)
        denied = self._put_json(
            "/api/audit",
            {
                "version": 3,
                "event_id": str(uuid.uuid4()),
                "audit_session_id": audit_session_id,
                "note": "tampered",
            },
            token=outsider_token,
        )
        self.assertEqual(denied.status_code, 404)
        allowed = self._put_json(
            "/api/audit",
            {
                "version": 3,
                "event_id": str(uuid.uuid4()),
                "audit_session_id": audit_session_id,
                "note": "approved",
            },
            token=controller_token,
        )
        self.assertEqual(allowed.status_code, 200, allowed.content)
        connection_log.refresh_from_db()
        self.assertEqual(connection_log.note, "approved")

        close_event = {
            "version": 3,
            "event_id": str(uuid.uuid4()),
            "audit_session_id": audit_session_id,
            "action": "close",
            "id": "111111111",
            "uuid": host_uuid,
            "conn_id": 7,
            "session_id": 99,
        }
        self.assertEqual(
            self._post_json("/api/audit/conn", close_event, token=host_token).status_code,
            200,
        )
        self.assertEqual(
            self._post_json("/api/audit/conn", close_event, token=host_token).status_code,
            200,
        )

        restarted_process_event = {
            **new_event,
            "event_id": str(uuid.uuid4()),
            "session_id": 100,
        }
        restarted = self._post_json(
            "/api/audit/conn",
            restarted_process_event,
            token=host_token,
        )
        self.assertEqual(restarted.status_code, 201, restarted.content)
        restarted_audit_session_id = restarted.json()["audit_session_id"]
        restarted_update = self._post_json(
            "/api/audit/conn",
            {
                "version": 3,
                "event_id": str(uuid.uuid4()),
                "audit_session_id": restarted_audit_session_id,
                "action": "update",
                "id": "111111111",
                "uuid": host_uuid,
                "conn_id": 7,
                "session_id": 100,
                "peer": ["222222222", "bob"],
                "type": 0,
                "primary_auth": 3,
                "two_factor": 1,
            },
            token=host_token,
        )
        self.assertEqual(restarted_update.status_code, 200, restarted_update.content)
        restarted_active = self.client.get(
            f"/api/audit/conn/active?version=3&event_id={uuid.uuid4()}&id=111111111&session_id=100&conn_type=0",
            **self._auth_headers(controller_token),
        )
        self.assertEqual(restarted_active.status_code, 200, restarted_active.content)
        self.assertEqual(restarted_active.json()["audit_session_id"], restarted_audit_session_id)
        self.assertEqual(ConnLog.objects.count(), 2)

        file_event = {
            "version": 4,
            "receipt_version": 1,
            "event_id": str(uuid.uuid4()),
            "reporter_sequence": ConnLog.objects.get(guid=restarted_audit_session_id).event_revision + 1,
            "audit_session_id": restarted_audit_session_id,
            "transfer_id": str(uuid.uuid4()),
            "transfer_revision": 1,
            "state": "started",
            "id": "111111111",
            "uuid": host_uuid,
            "peer_id": "222222222",
            "conn_id": 7,
            "direction": 0,
            "path": "/documents",
            "is_file": False,
            "planned_file_count": 1,
            "planned_bytes": 4096,
            "transferred_bytes": 0,
            "sample_files": [{"path": "report.pdf", "size": 4096}],
            "source_kind": "file_transfer",
            "terminal_reason": "",
        }
        self.assertEqual(
            self._post_json("/api/audit/file", file_event, token=host_token).status_code,
            200,
        )
        file_log = FileLog.objects.get()
        self.assertEqual(file_log.filesize, 4096)
        self.assertEqual(file_log.planned_bytes, 4096)
        self.assertEqual(file_log.sample_files, [{"path": "report.pdf", "size": 4096}])

        alarm_event = {
            "version": 3,
            "receipt_version": 1,
            "event_id": str(uuid.uuid4()),
            "reporter_sequence": ConnLog.objects.get(guid=restarted_audit_session_id).event_revision + 1,
            "audit_session_id": restarted_audit_session_id,
            "id": "111111111",
            "uuid": host_uuid,
            "typ": AlarmLog.TYPE_SESSION_SCOPE_VIOLATION,
            "conn_id": 7,
            "info": json.dumps({"message": "scope mismatch"}),
        }
        self.assertEqual(
            self._post_json("/api/audit/alarm", alarm_event, token=host_token).status_code,
            200,
        )
        alarm = AlarmLog.objects.get()
        self.assertEqual(alarm.reporter_device_id, "111111111")
        self.assertEqual(alarm.info, {"message": "scope mismatch"})


class OperationalEndpointTests(TestCase):
    def setUp(self):
        self.recording_root = tempfile.TemporaryDirectory()
        self.recording_settings = override_settings(
            RECORD_UPLOAD_ROOT=Path(self.recording_root.name),
            RECORD_UPLOAD_REQUIRE_MOUNT=False,
            RECORD_UPLOAD_VOLUME_RESERVE_BYTES=0,
            RECORD_UPLOAD_VOLUME_RESERVE_INODES=0,
            RECORD_UPLOAD_CAPABILITY_CACHE_SECONDS=0,
        )
        self.recording_settings.enable()

    def tearDown(self):
        self.recording_settings.disable()
        self.recording_root.cleanup()

    @override_settings(
        DEVICE_VERIFICATION_TOKEN="v" * 48,
        SECURE_SSL_REDIRECT=False,
        RECORD_UPLOAD_REQUIRE_MOUNT=False,
    )
    def test_readiness_rejects_a_missing_runtime_recording_volume(self):
        with tempfile.TemporaryDirectory() as parent:
            missing_root = Path(parent) / "removed-recording-volume"
            with override_settings(RECORD_UPLOAD_ROOT=missing_root):
                response = self.client.get("/health/ready")

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.json(), {"status": "not_ready"})

    @override_settings(
        DEVICE_VERIFICATION_TOKEN="v" * 48,
        SECURE_SSL_REDIRECT=False,
    )
    def test_readiness_rejects_a_wrong_but_well_formed_data_key(self):
        pending = OidcPendingAuth.objects.create(
            state="readiness-key-check",
            poll_code_hash=hashlib.sha256(b"readiness-poll-code").hexdigest(),
            provider="example",
            nonce="readiness-nonce",
            code_verifier="readiness-code-verifier",
        )
        wrong_key = b"w" * 32

        with override_settings(
            DATA_ENCRYPTION_KEY_BYTES=wrong_key,
            DATA_ENCRYPTION_PRIMARY_KEY_ID="wrong-key",
            DATA_ENCRYPTION_KEYS={"wrong-key": wrong_key},
            DATA_ENCRYPTION_V1_KEY_ID="wrong-key",
        ):
            with self.assertRaises(ValidationError):
                pending.refresh_from_db()
            response = self.client.get("/health/ready")

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.json(), {"status": "not_ready"})

    @override_settings(
        DEVICE_VERIFICATION_TOKEN="v" * 48,
        SECURE_PROXY_SSL_HEADER=("HTTP_X_FORWARDED_PROTO", "https"),
        SECURE_SSL_REDIRECT=True,
    )
    def test_tls_readiness_probe_requires_trusted_https_marker(self):
        insecure = self.client.get("/health/ready")
        self.assertEqual(insecure.status_code, 301)
        self.assertTrue(insecure["Location"].startswith("https://"))

        secure = self.client.get(
            "/health/ready",
            HTTP_X_FORWARDED_PROTO="https",
        )
        self.assertEqual(secure.status_code, 200)
        self.assertEqual(secure.json(), {"status": "ready"})

    @override_settings(
        DEVICE_VERIFICATION_TOKEN="",
        SECURE_SSL_REDIRECT=False,
    )
    def test_readiness_rejects_missing_device_verification_secret(self):
        response = self.client.get("/health/ready")

        self.assertEqual(response.status_code, 503)
        self.assertEqual(response.json(), {"status": "not_ready"})

    @override_settings(SECURE_SSL_REDIRECT=False)
    def test_liveness_does_not_depend_on_database_readiness(self):
        response = self.client.get("/health/live")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json(), {"status": "live"})


class DataEncryptionRotationTests(TestCase):
    def test_rotation_rewraps_recording_data_keys_before_retiring_the_old_kek(self):
        encoded_data_key = recording_crypto.encode_data_key(b"d" * recording_crypto.DATA_KEY_BYTES)
        storage_object_id = uuid.uuid4()
        old_key_id = project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID
        upload = RecordingUpload.objects.create(
            create_id=uuid.uuid4(),
            device_id=None,
            device_id_at_create=1,
            owner_id_at_create=1,
            deployment_generation=0,
            device_rid_at_create="rotation-device",
            device_uuid_at_create="rotation-device-uuid",
            storage_object_id=storage_object_id,
            storage_namespace=ingestion_governance.recording_namespace(storage_object_id),
            filename="rotation-recording.webm",
            encryption_version=recording_crypto.FORMAT_VERSION,
            data_key_kek_id=old_key_id,
            encrypted_data_key=encoded_data_key,
            storage_offset=recording_crypto.HEADER_SIZE,
        )
        database_upload_id = RecordingUpload._meta.pk.get_db_prep_value(upload.pk, connection)
        old_key = project_settings.DATA_ENCRYPTION_KEYS[old_key_id]
        new_key_id = "recording-rotation"
        new_key = b"n" * 32
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT encrypted_data_key FROM api_recordingupload WHERE upload_id = %s",
                [database_upload_id],
            )
            old_envelope = cursor.fetchone()[0]
        self.assertTrue(old_envelope.startswith(f"{FIELD_PREFIX}{old_key_id}:"))
        self.assertNotIn(encoded_data_key, old_envelope)

        with override_settings(
            DATA_ENCRYPTION_KEY_BYTES=new_key,
            DATA_ENCRYPTION_KEYS={old_key_id: old_key, new_key_id: new_key},
            DATA_ENCRYPTION_PRIMARY_KEY_ID=new_key_id,
            DATA_ENCRYPTION_V1_KEY_ID=old_key_id,
        ):
            output = StringIO()
            call_command("rotate_data_encryption", stdout=output)
            self.assertEqual(json.loads(output.getvalue())["rewritten"], 1)
            upload.refresh_from_db()
            self.assertEqual(upload.encrypted_data_key, encoded_data_key)
            self.assertEqual(upload.data_key_kek_id, new_key_id)
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT encrypted_data_key FROM api_recordingupload WHERE upload_id = %s",
                    [database_upload_id],
                )
                new_envelope = cursor.fetchone()[0]
            self.assertTrue(new_envelope.startswith(f"{FIELD_PREFIX}{new_key_id}:"))
            self.assertNotIn(encoded_data_key, new_envelope)
            call_command(
                "rotate_data_encryption",
                retire_key_id=old_key_id,
                stdout=StringIO(),
            )
            self.assertFalse(DataEncryptionKeyState.objects.filter(key_id=old_key_id).exists())

    def test_rotation_is_resumable_and_legacy_key_retirement_is_explicit(self):
        recording_root = self.enterContext(tempfile.TemporaryDirectory())
        old_key_id = project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID
        old_key = project_settings.DATA_ENCRYPTION_KEYS[old_key_id]
        pending = OidcPendingAuth.objects.create(
            state="rotation-state",
            poll_code_hash=hashlib.sha256(b"rotation-poll-code").hexdigest(),
            provider="example",
            nonce="legacy-v1-nonce",
            code_verifier="old-v2-verifier",
        )
        legacy_ciphertext = SecretBox(old_key).encrypt(b"legacy-v1-nonce")
        legacy_envelope = "secretbox:v1:" + base64.b64encode(bytes(legacy_ciphertext)).decode()
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE api_oidcpendingauth SET nonce = %s WHERE state = %s",
                [legacy_envelope, pending.pk],
            )

        new_key_id = "rotation-2026-08"
        new_key = b"n" * 32
        rotated_prefix = f"{FIELD_PREFIX}{new_key_id}:"
        rotation_settings = {
            "DATA_ENCRYPTION_KEY_BYTES": new_key,
            "DATA_ENCRYPTION_KEYS": {old_key_id: old_key, new_key_id: new_key},
            "DATA_ENCRYPTION_PRIMARY_KEY_ID": new_key_id,
            "DATA_ENCRYPTION_V1_KEY_ID": old_key_id,
            "DEVICE_VERIFICATION_TOKEN": "v" * 48,
            "SECURE_SSL_REDIRECT": False,
            "RECORD_UPLOAD_ROOT": Path(recording_root),
            "RECORD_UPLOAD_REQUIRE_MOUNT": False,
            "RECORD_UPLOAD_VOLUME_RESERVE_BYTES": 0,
            "RECORD_UPLOAD_VOLUME_RESERVE_INODES": 0,
            "RECORD_UPLOAD_CAPABILITY_CACHE_SECONDS": 0,
        }
        with override_settings(**rotation_settings):
            first_output = StringIO()
            call_command(
                "rotate_data_encryption",
                batch_size=1,
                max_batches=1,
                stdout=first_output,
            )
            first_result = json.loads(first_output.getvalue())
            self.assertEqual(first_result["rewritten"], 1)
            self.assertEqual(first_result["batches"], 1)

            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT nonce, code_verifier FROM api_oidcpendingauth WHERE state = %s",
                    [pending.pk],
                )
                first_nonce, first_verifier = cursor.fetchone()
            self.assertTrue(first_nonce.startswith(rotated_prefix))
            self.assertFalse(first_verifier.startswith(rotated_prefix))

            call_command("rotate_data_encryption", batch_size=1, stdout=StringIO())
            pending.refresh_from_db()
            self.assertEqual(pending.nonce, "legacy-v1-nonce")
            self.assertEqual(pending.code_verifier, "old-v2-verifier")
            with connection.cursor() as cursor:
                cursor.execute(
                    "SELECT nonce, code_verifier FROM api_oidcpendingauth WHERE state = %s",
                    [pending.pk],
                )
                rotated_nonce, rotated_verifier = cursor.fetchone()
            self.assertTrue(rotated_nonce.startswith(rotated_prefix))
            self.assertTrue(rotated_verifier.startswith(rotated_prefix))
            self.assertEqual(
                list(DataEncryptionKeyState.objects.filter(is_primary=True).values_list("key_id", flat=True)),
                [new_key_id],
            )
            self.assertEqual(self.client.get("/health/ready").status_code, 200)

            call_command(
                "rotate_data_encryption",
                retire_key_id=old_key_id,
                stdout=StringIO(),
            )
            self.assertFalse(DataEncryptionKeyState.objects.filter(key_id=old_key_id).exists())

    def test_rotation_does_not_retire_a_referenced_key(self):
        old_key_id = project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID
        old_key = project_settings.DATA_ENCRYPTION_KEYS[old_key_id]
        OidcPendingAuth.objects.create(
            state="referenced-key-state",
            poll_code_hash=hashlib.sha256(b"referenced-key-poll-code").hexdigest(),
            provider="example",
            nonce="nonce",
            code_verifier="verifier",
        )
        new_key_id = "retirement-candidate"
        new_key = b"r" * 32
        with (
            override_settings(
                DATA_ENCRYPTION_KEY_BYTES=new_key,
                DATA_ENCRYPTION_KEYS={old_key_id: old_key, new_key_id: new_key},
                DATA_ENCRYPTION_PRIMARY_KEY_ID=new_key_id,
                DATA_ENCRYPTION_V1_KEY_ID=old_key_id,
            ),
            self.assertRaisesMessage(CommandError, "still referenced"),
        ):
            call_command(
                "rotate_data_encryption",
                max_batches=1,
                retire_key_id=old_key_id,
                stdout=StringIO(),
            )

    def test_bounded_rotation_makes_progress_across_repeated_invocations(self):
        old_key_id = project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID
        old_key = project_settings.DATA_ENCRYPTION_KEYS[old_key_id]
        for index in range(2):
            OidcPendingAuth.objects.create(
                state=f"bounded-rotation-{index}",
                poll_code_hash=hashlib.sha256(f"bounded-poll-{index}".encode()).hexdigest(),
                provider="example",
                nonce=f"nonce-{index}",
                code_verifier=f"verifier-{index}",
            )
        new_key_id = "bounded-primary"
        new_key = b"b" * 32
        with override_settings(
            DATA_ENCRYPTION_KEY_BYTES=new_key,
            DATA_ENCRYPTION_KEYS={old_key_id: old_key, new_key_id: new_key},
            DATA_ENCRYPTION_PRIMARY_KEY_ID=new_key_id,
            DATA_ENCRYPTION_V1_KEY_ID=old_key_id,
        ):
            remaining = []
            for _index in range(4):
                output = StringIO()
                call_command(
                    "rotate_data_encryption",
                    batch_size=1,
                    max_batches=1,
                    stdout=output,
                )
                result = json.loads(output.getvalue())
                self.assertEqual(result["rewritten"], 1)
                remaining.append(result["remaining"])

        self.assertEqual(remaining, [3, 2, 1, 0])

    def test_full_rotation_authenticates_existing_primary_envelopes(self):
        pending = OidcPendingAuth.objects.create(
            state="corrupt-primary-envelope",
            poll_code_hash=hashlib.sha256(b"corrupt-primary-poll").hexdigest(),
            provider="example",
            nonce="nonce-to-corrupt",
            code_verifier="valid-verifier",
        )
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT nonce FROM api_oidcpendingauth WHERE state = %s",
                [pending.pk],
            )
            envelope = cursor.fetchone()[0]
            prefix, encoded = envelope.rsplit(":", 1)
            ciphertext = bytearray(base64.b64decode(encoded))
            ciphertext[-1] ^= 1
            cursor.execute(
                "UPDATE api_oidcpendingauth SET nonce = %s WHERE state = %s",
                [f"{prefix}:{base64.b64encode(ciphertext).decode()}", pending.pk],
            )

        with self.assertRaisesMessage(CommandError, "authentication failed"):
            call_command("rotate_data_encryption", stdout=StringIO())

    def test_key_id_like_wildcards_do_not_hide_non_primary_values(self):
        primary_key_id = "key_a"
        confusing_key_id = "keyxa"
        primary_key = b"p" * 32
        confusing_key = b"x" * 32
        default_key_id = project_settings.DATA_ENCRYPTION_PRIMARY_KEY_ID
        default_key = project_settings.DATA_ENCRYPTION_KEYS[default_key_id]
        keyring = {
            default_key_id: default_key,
            primary_key_id: primary_key,
            confusing_key_id: confusing_key,
        }
        with override_settings(
            DATA_ENCRYPTION_KEY_BYTES=primary_key,
            DATA_ENCRYPTION_KEYS=keyring,
            DATA_ENCRYPTION_PRIMARY_KEY_ID=primary_key_id,
            DATA_ENCRYPTION_V1_KEY_ID=default_key_id,
        ):
            DataEncryptionKeyState.objects.create(
                key_id=confusing_key_id,
                key_fingerprint=key_fingerprint(confusing_key),
                encrypted_canary=key_canary(confusing_key_id),
                is_primary=False,
            )
            pending = OidcPendingAuth.objects.create(
                state="wildcard-key-id",
                poll_code_hash=hashlib.sha256(b"wildcard-key-poll").hexdigest(),
                provider="example",
                nonce="primary-value",
                code_verifier="primary-verifier",
            )
            with connection.cursor() as cursor:
                cursor.execute(
                    "UPDATE api_oidcpendingauth SET nonce = %s WHERE state = %s",
                    [encrypt_text("confusing-value", key_id=confusing_key_id), pending.pk],
                )
            output = StringIO()
            call_command("rotate_data_encryption", dry_run=True, stdout=output)

        self.assertEqual(json.loads(output.getvalue())["remaining"], 1)
